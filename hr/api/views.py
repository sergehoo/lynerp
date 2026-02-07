# Lyneerp/hr/views.py
import csv
import io
import logging
import uuid
from datetime import datetime
from typing import Dict, Any, List, Optional
import pandas as pd
from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.cache import cache
# from django.contrib.auth.models import User
from django.http import HttpResponse, HttpRequest
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.db.models import Count, Avg, Q
from django.db import transaction, IntegrityError, models
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from rest_framework import viewsets, status, filters, permissions, serializers
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import IsAuthenticated, BasePermission, SAFE_METHODS
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, JSONParser
from rest_framework.views import APIView
from hr.ai_recruitment_service import AIRecruitmentService
from hr.permissions import HasRHAccess, HasRole

# Models
from hr.models import (
    Department,
    Employee,
    LeaveRequest,
    LeaveType,
    LeaveBalance,
    Position,
    Attendance,
    PerformanceReview,
    JobApplication,
    Recruitment,
    Interview, ContractType, EmploymentContract, ContractAmendment, ContractTemplate, ContractAlert, ContractHistory,
    SalaryHistory, HRDocument, LeaveApprovalStep, HolidayCalendar, Holiday, WorkScheduleTemplate, MedicalRecord,
    MedicalVisit, MedicalRestriction, Payroll, RecruitmentAnalytics, JobOffer, RecruitmentWorkflow, InterviewFeedback,
)

# Serializers
from hr.api.serializers import (
    DepartmentSerializer,
    EmployeeSerializer,
    LeaveRequestSerializer,
    LeaveTypeSerializer,
    LeaveBalanceSerializer,
    PositionSerializer,
    AttendanceSerializer,
    PerformanceReviewSerializer,
    JobApplicationSerializer,
    JobApplicationDetailSerializer,
    InterviewSerializer,
    AIProcessingResultSerializer,
    BulkLeaveActionSerializer,
    RecruitmentStatsSerializer,
    HRDashboardSerializer,
    BulkEmployeeActionSerializer,
    EmployeeFilterSerializer,
    EmployeeImportSerializer,
    EmployeeExportSerializer,
    LeaveRequestFilterSerializer,
    EmployeeStatsSerializer,
    RecruitmentSerializer,
    RecruitmentFilterSerializer,
    AIProcessingResult, TenantLiteSerializer, LeaveApprovalStepSerializer, ContractTypeSerializer,
    EmploymentContractSerializer, ContractAmendmentSerializer, ContractTemplateSerializer, ContractAlertSerializer,
    ContractHistorySerializer, SalaryHistorySerializer, HRDocumentSerializer, HolidayCalendarSerializer,
    HolidaySerializer, WorkScheduleTemplateSerializer, MedicalRecordSerializer, MedicalVisitSerializer,
    MedicalRestrictionSerializer, PayrollSerializer, RecruitmentAnalyticsSerializer, JobOfferSerializer,
    RecruitmentWorkflowSerializer, InterviewFeedbackSerializer, EmploymentContractExportSerializer,
)
from tenants.models import Tenant, TenantDomain, TenantUser

# Services (export, etc.)
try:
    from ..services import EmployeeExportService
except Exception:
    # Fallback minimal si le service n'est pas encore impl√©ment√©
    class EmployeeExportService:
        def export_employees(self, tenant_id: str, export_format: str, fields: List[str], filters: Dict[str, Any]):
            return {"success": False, "error": "EmployeeExportService non impl√©ment√©"}

logger = logging.getLogger(__name__)


def get_current_tenant_from_request(request: HttpRequest) -> Optional[Tenant]:
    """
    R√©sout le tenant √† partir, dans l'ordre :
    1) du token OIDC (request.oidc.tenant / tenant_id : slug ou UUID)
    2) du header X-Tenant-Id : slug ou UUID
    3) du host (TenantDomain ou sous-domaine = slug)
    """

    # 1) Via OIDC (si tu ajoutes oidc au request)
    oidc = getattr(request, "oidc", {}) or {}
    raw = oidc.get("tenant") or oidc.get("tenant_id")
    if raw:
        # slug
        t = Tenant.objects.filter(slug=raw).first()
        if t:
            return t
        # UUID
        try:
            uuid_val = uuid.UUID(str(raw))
            t = Tenant.objects.filter(id=uuid_val).first()
            if t:
                return t
        except ValueError:
            pass

    # 2) Header X-Tenant-Id
    hdr = request.headers.get("X-Tenant-Id") or request.META.get("HTTP_X_TENANT_ID")
    if hdr:
        t = Tenant.objects.filter(slug=hdr).first()
        if t:
            return t
        try:
            uuid_val = uuid.UUID(str(hdr))
            t = Tenant.objects.filter(id=uuid_val).first()
            if t:
                return t
        except ValueError:
            pass

    # 3) Par le host
    host = request.get_host().split(":")[0].lower()

    # 3.a) TenantDomain direct
    dom = TenantDomain.objects.filter(domain=host).select_related("tenant").first()
    if dom:
        return dom.tenant

    # 3.b) Sous-domaine -> slug
    parts = host.split(".")
    if len(parts) >= 3:
        sub = parts[0]
        t = Tenant.objects.filter(slug=sub).first()
        if t:
            return t

    return None


# -----------------------------
# Mixins multi-tenant
# -----------------------------
class BaseTenantViewSet(viewsets.ModelViewSet):
    TENANT_HEADER = "X-Tenant-Id"

    def get_tenant_id(self):
        """
        Renvoie tenant_id depuis:
        - request.tenant_id (middleware)
        - request.tenant (middleware)
        - Header X-Tenant-Id
        """
        # 1) middleware qui pose request.tenant_id
        tenant_id = getattr(self.request, "tenant_id", None)
        if tenant_id:
            return str(tenant_id)

        # 2) middleware qui pose request.tenant
        tenant = getattr(self.request, "tenant", None)
        if tenant:
            # slug sinon id
            return str(getattr(tenant, "slug", None) or tenant.id)

        # 3) header
        return self.request.headers.get(self.TENANT_HEADER)

    def get_tenant(self):
        """
        Retourne l'objet Tenant (ou None).
        """
        tenant = getattr(self.request, "tenant", None)
        if tenant:
            return tenant

        tenant_id = self.get_tenant_id()
        if not tenant_id:
            return None

        # accepte id OU slug
        return (
            Tenant.objects
            .filter(Q(id=tenant_id) | Q(slug=tenant_id))
            .first()
        )

    def get_queryset(self):
        # 0) Super admin plateforme : acc√®s global
        if self.request.user and self.request.user.is_superuser:
            return super().get_queryset()

        tenant = self.get_tenant()
        if not tenant:
            return self.queryset.none()

        qs = super().get_queryset()
        model = qs.model

        # FK tenant
        if hasattr(model, "tenant"):
            return qs.filter(tenant=tenant)

        # tenant_id string
        if hasattr(model, "tenant_id"):
            slug = getattr(tenant, "slug", None)
            filt = Q(tenant_id=str(tenant.id))
            if slug:
                filt |= Q(tenant_id=slug)
            return qs.filter(filt).distinct()

        return qs.none()


def get_membership(user, tenant: Tenant) -> Optional[TenantUser]:
    if not user or user.is_anonymous or not tenant:
        return None
    return (
        TenantUser.objects
        .filter(user=user, tenant=tenant, is_active=True)
        .select_related("tenant", "user")
        .first()
    )


class HasTenantAccess(BasePermission):
    """
    Autorise:
    - superuser: tout
    - tenant ADMIN/OWNER/HR_BPO selon action
    """
    allowed_roles_read = {"OWNER", "ADMIN", "MANAGER", "HR_BPO", "VIEWER", "MEMBER"}
    allowed_roles_write = {"OWNER", "ADMIN", "HR_BPO"}  # RH externalis√©e peut cr√©er/modifier RH

    def has_permission(self, request, view):
        if request.user and request.user.is_superuser:
            return True

        tenant = getattr(view, "get_tenant", lambda: None)()
        # si ton viewset n'a pas get_tenant, tu peux faire:
        # tenant = get_current_tenant_from_request(request)

        if not tenant:
            return False

        membership = get_membership(request.user, tenant)
        if not membership:
            return False

        if request.method in SAFE_METHODS:
            return membership.role in self.allowed_roles_read
        return membership.role in self.allowed_roles_write


class TenantViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Tenant.objects.filter(is_active=True).order_by("name")
    serializer_class = TenantLiteSerializer
    permission_classes = [IsAuthenticated]


# -----------------------------
# Dashboard RH
# -----------------------------

class HRDashboardViewSet(viewsets.ViewSet):
    """Vues pour le tableau de bord RH"""
    permission_classes = [IsAuthenticated, HasRHAccess]

    # -----------------------------
    # Tenant helpers
    # -----------------------------
    def get_tenant(self, request) -> Optional[Tenant]:
        return get_current_tenant_from_request(request)

    def tenant_filter_for(self, model_cls, tenant: Tenant) -> Q:
        """
        Retourne un Q adapt√© selon le type de stockage du tenant sur model_cls:
        - tenant = FK -> tenant=tenant
        - tenant_id = UUIDField -> tenant_id=tenant.id
        - tenant_id = CharField/TextField -> tenant_id in (tenant.slug, str(tenant.id))
        """
        # 1) FK tenant
        if any(f.name == "tenant" for f in model_cls._meta.fields):
            return Q(tenant=tenant)

        # 2) champ tenant_id
        try:
            f = model_cls._meta.get_field("tenant_id")
        except Exception:
            # Aucun champ tenant reconnu -> pas de fuite de donn√©es
            return Q(pk__in=[])

        if isinstance(f, models.UUIDField):
            return Q(tenant_id=tenant.id)

        # CharField / TextField / autres => compat
        return Q(tenant_id=tenant.slug) | Q(tenant_id=str(tenant.id))

    def filter_by_tenant(self, qs, model_cls, tenant: Tenant):
        """
        Applique un filtre tenant sur un queryset, en fallback safe.
        """
        if not tenant:
            return qs.none()
        q = self.tenant_filter_for(model_cls, tenant)
        if not q.children:
            return qs.none()
        return qs.filter(q)

    # -----------------------------
    # Dashboard: Stats principales
    # -----------------------------
    @action(detail=False, methods=["get"])
    def stats(self, request):
        tenant = self.get_tenant(request)
        if not tenant:
            return Response(
                {"detail": "Tenant introuvable pour cette requ√™te"},
                status=status.HTTP_400_BAD_REQUEST
            )

        cache_key = f"hr:dash:stats:{tenant.id}"
        cached = cache.get(cache_key)
        if cached:
            return Response(cached)

        today = timezone.localdate()

        # Models avec tenant = FK(Tenant)
        emp_qs = Employee.objects.filter(tenant=tenant)
        total_employees = emp_qs.count()
        active_employees = emp_qs.filter(is_active=True).count()

        employees_on_leave = (
            emp_qs.filter(
                is_active=True,
                leaverequest__status="approved",
                leaverequest__start_date__lte=today,
                leaverequest__end_date__gte=today,
            )
            .distinct()
            .count()
        )

        new_hires_this_month = emp_qs.filter(
            hire_date__year=today.year,
            hire_date__month=today.month,
        ).count()

        # Models qui peuvent stocker tenant_id
        pending_leave_requests = self.filter_by_tenant(
            LeaveRequest.objects.all(), LeaveRequest, tenant
        ).filter(status="pending").count()

        active_recruitments = self.filter_by_tenant(
            Recruitment.objects.all(), Recruitment, tenant
        ).filter(status__in=["OPEN", "IN_REVIEW", "INTERVIEW", "OFFER"]).count()

        upcoming_reviews = self.filter_by_tenant(
            PerformanceReview.objects.all(), PerformanceReview, tenant
        ).filter(review_date__gte=today, status="DRAFT").count()

        stats_data = {
            "total_employees": total_employees,
            "active_employees": active_employees,
            "employees_on_leave": employees_on_leave,
            "new_hires_this_month": new_hires_this_month,
            "pending_leave_requests": pending_leave_requests,
            "active_recruitments": active_recruitments,
            "upcoming_reviews": upcoming_reviews,
        }

        data = HRDashboardSerializer(stats_data).data
        cache.set(cache_key, data, 30)
        return Response(data)

    # -----------------------------
    # Dashboard: listes l√©g√®res
    # -----------------------------
    @action(detail=False, methods=["get"])
    def latest_hires(self, request):
        tenant = self.get_tenant(request)
        if not tenant:
            return Response([], status=200)

        try:
            limit = min(int(request.query_params.get("limit", 5)), 50)
        except Exception:
            limit = 5

        qs = (
            Employee.objects
            .filter(tenant=tenant)
            .select_related("department", "position")
            .only("id", "first_name", "last_name", "hire_date", "department__name", "position__title")
            .order_by("-hire_date")[:limit]
        )

        data = [{
            "id": str(e.id),
            "first_name": e.first_name,
            "last_name": e.last_name,
            "hire_date": e.hire_date,
            "department_name": getattr(e.department, "name", None),
            "position_title": getattr(e.position, "title", None),
        } for e in qs]

        return Response(data)

    @action(detail=False, methods=["get"])
    def active_recruitments_list(self, request):
        tenant = self.get_tenant(request)
        if not tenant:
            return Response([], status=200)

        try:
            limit = min(int(request.query_params.get("limit", 3)), 50)
        except Exception:
            limit = 3

        qs = self.filter_by_tenant(
            Recruitment.objects.all(), Recruitment, tenant
        ).filter(status__in=["OPEN", "IN_REVIEW", "INTERVIEW", "OFFER"])

        qs = (
            qs.select_related("department", "position")
            .only(
                "id", "title", "status", "publication_date", "number_of_positions",
                "department__name", "position__title"
            )
            .order_by("-publication_date", "-created_at")[:limit]
        )

        data = [{
            "id": r.id,
            "title": r.title,
            "status": r.status,
            "publication_date": r.publication_date,
            "number_of_positions": r.number_of_positions,
            "department_name": getattr(r.department, "name", None),
            "position_title": getattr(r.position, "title", None),
        } for r in qs]

        return Response(data)

    # -----------------------------
    # Statistiques Recrutement
    # -----------------------------
    @action(detail=False, methods=["get"])
    def recruitment_stats(self, request):
        tenant = self.get_tenant(request)
        if not tenant:
            return Response({"detail": "Tenant introuvable"}, status=400)

        cache_key = f"hr:dash:recruitment_stats:{tenant.id}"
        cached = cache.get(cache_key)
        if cached:
            return Response(cached)

        recruit_qs = self.filter_by_tenant(Recruitment.objects.all(), Recruitment, tenant)
        app_qs = self.filter_by_tenant(JobApplication.objects.all(), JobApplication, tenant)
        ai_qs = self.filter_by_tenant(AIProcessingResult.objects.all(), AIProcessingResult, tenant)

        total_recruitments = recruit_qs.count()
        active_recruitments = recruit_qs.filter(status__in=["OPEN", "IN_REVIEW", "INTERVIEW", "OFFER"]).count()

        total_applications = app_qs.count()
        applications_this_week = app_qs.filter(
            applied_at__gte=timezone.now() - timezone.timedelta(days=7)
        ).count()

        avg_ai_score = (
            app_qs.filter(ai_score__isnull=False)
            .aggregate(avg_score=Avg("ai_score"))["avg_score"]
            or 0
        )

        apps_by_status = dict(
            app_qs.values("status")
            .annotate(count=Count("id"))
            .values_list("status", "count")
        )

        hires = app_qs.filter(status="HIRED").count()
        hire_conversion_rate = (hires / total_applications * 100.0) if total_applications > 0 else 0.0

        ai_completed = ai_qs.filter(status="COMPLETED").count()
        ai_failed = ai_qs.filter(status="FAILED").count()
        ai_avg_overall = ai_qs.aggregate(a=Avg("overall_match_score"))["a"] or 0.0

        stats_data = {
            "total_recruitments": total_recruitments,
            "active_recruitments": active_recruitments,
            "total_applications": total_applications,
            "applications_this_week": applications_this_week,
            "average_ai_score": round(float(avg_ai_score), 2),
            "hire_conversion_rate": round(float(hire_conversion_rate), 2),
            "average_time_to_hire": 0.0,  # TODO
            "applications_by_status": apps_by_status,
            "ai_processing_stats": {
                "completed": ai_completed,
                "failed": ai_failed,
                "avg_overall_match_score": round(float(ai_avg_overall), 2),
            },
        }

        data = RecruitmentStatsSerializer(stats_data).data
        cache.set(cache_key, data, 30)
        return Response(data)

    # -----------------------------
    # Stats Employ√©s
    # -----------------------------
    @action(detail=False, methods=["get"])
    def employee_stats(self, request):
        tenant = self.get_tenant(request)
        if not tenant:
            return Response({"detail": "Tenant introuvable"}, status=400)

        by_department = dict(
            Employee.objects
            .filter(tenant=tenant, is_active=True)
            .values("department__name")
            .annotate(count=Count("id"))
            .values_list("department__name", "count")
        )

        by_contract = dict(
            Employee.objects
            .filter(tenant=tenant, is_active=True)
            .values("contract_type")
            .annotate(count=Count("id"))
            .values_list("contract_type", "count")
        )

        gender_dist = dict(
            Employee.objects
            .filter(tenant=tenant, is_active=True)
            .exclude(gender="")
            .values("gender")
            .annotate(count=Count("id"))
            .values_list("gender", "count")
        )

        stats_data = {
            "total_by_department": by_department,
            "total_by_contract_type": by_contract,
            "gender_distribution": gender_dist,
        }

        return Response(EmployeeStatsSerializer(stats_data).data)


# -----------------------------
# Actions en masse
# -----------------------------
class BulkActionsViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated, HasRHAccess]

    @action(detail=False, methods=['post'])
    def bulk_leave_action(self, request):
        """Action batch sur les demandes de cong√©"""
        tenant = get_current_tenant_from_request(request)  # üëà
        serializer = BulkLeaveActionSerializer(data=request.data)
        if serializer.is_valid():
            leave_request_ids = serializer.validated_data['leave_request_ids']
            action_type = serializer.validated_data['action']
            reason = serializer.validated_data.get('reason', '')

            leave_requests = LeaveRequest.objects.filter(
                id__in=leave_request_ids,
                tenant_id=tenant.slug,  # üëà coh√©rent avec CharField
            )

            updated_count = 0
            with transaction.atomic():
                for leave_request in leave_requests.select_for_update():
                    if action_type == 'approve':
                        leave_request.status = 'approved'
                        leave_request.approved_by = getattr(request.user, "employee_profile", None)
                        leave_request.approved_at = timezone.now()
                    elif action_type == 'reject':
                        leave_request.status = 'rejected'
                        leave_request.rejection_reason = reason
                        leave_request.approved_by = getattr(request.user, "employee_profile", None)
                        leave_request.approved_at = timezone.now()
                    elif action_type == 'cancel':
                        leave_request.status = 'cancelled'

                    leave_request.save()
                    updated_count += 1

            return Response({
                "message": f"{updated_count} demandes de cong√© mises √† jour",
                "action": action_type
            })
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], url_path='bulk_employee_action')
    def bulk_employee_action(self, request):
        tenant = get_current_tenant_from_request(request)
        if not tenant:
            return Response({"detail": "Tenant introuvable"}, status=400)

        # payload attendu: { employee_ids: [...], action: 'activate'|'deactivate'|'change_department'|'terminate', data: {...} }
        employee_ids = request.data.get("employee_ids") or []
        action_type = request.data.get("action")
        data = request.data.get("data") or {}

        if not employee_ids or not action_type:
            return Response({"detail": "employee_ids et action sont requis"}, status=400)

        qs = Employee.objects.filter(id__in=employee_ids, tenant=tenant)

        updated = 0
        with transaction.atomic():
            if action_type == "activate":
                updated = qs.update(is_active=True)

            elif action_type == "deactivate":
                updated = qs.update(is_active=False)

            elif action_type == "change_department":
                dept_id = data.get("department_id")
                if not dept_id:
                    return Response({"detail": "department_id requis"}, status=400)
                updated = qs.update(department_id=dept_id)

            elif action_type == "terminate":
                # adapte √† ton mod√®le: contract_end_date/status/etc.
                reason = data.get("reason", "")
                # exemple simple:
                updated = qs.update(is_active=False)

            else:
                return Response({"detail": f"Action inconnue: {action_type}"}, status=400)

        return Response({"message": f"{updated} employ√©s mis √† jour", "action": action_type})


# -----------------------------
# ViewSets RH
# -----------------------------
class DepartmentViewSet(BaseTenantViewSet, viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated]
    # permission_classes = [IsAuthenticated, HasRHAccess]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'code']
    ordering_fields = ['name', 'created_at']

    @action(detail=True, methods=['get'])
    def employees(self, request, pk=None):
        """Liste des employ√©s du d√©partement"""
        department = self.get_object()
        employees = department.employee_set.filter(is_active=True)
        serializer = EmployeeSerializer(employees, many=True)
        return Response(serializer.data)


# class EmployeeViewSet(BaseTenantViewSet, viewsets.ModelViewSet):
#     queryset = Employee.objects.all()
#     serializer_class = EmployeeSerializer
#     permission_classes = [IsAuthenticated]
#     # permission_classes = [IsAuthenticated, HasRHAccess, HasRole]
#     # required_roles = ["hr:view"]
#     filter_backends = [filters.SearchFilter, filters.OrderingFilter]
#     search_fields = ['first_name', 'last_name', 'email', 'matricule']
#     ordering_fields = ['first_name', 'last_name', 'hire_date', 'created_at']
#
#     def get_queryset(self):
#         queryset = super().get_queryset()
#
#         # Filtrage avanc√©
#         filter_serializer = EmployeeFilterSerializer(data=self.request.query_params)
#         if filter_serializer.is_valid():
#             filt: Dict[str, Any] = {}
#             if filter_serializer.validated_data.get('department'):
#                 filt['department__name'] = filter_serializer.validated_data['department']
#             if filter_serializer.validated_data.get('position'):
#                 filt['position__title'] = filter_serializer.validated_data['position']
#             if filter_serializer.validated_data.get('contract_type'):
#                 filt['contract_type'] = filter_serializer.validated_data['contract_type']
#             if filter_serializer.validated_data.get('is_active') is not None:
#                 filt['is_active'] = filter_serializer.validated_data['is_active']
#             if filter_serializer.validated_data.get('hire_date_from'):
#                 filt['hire_date__gte'] = filter_serializer.validated_data['hire_date_from']
#             if filter_serializer.validated_data.get('hire_date_to'):
#                 filt['hire_date__lte'] = filter_serializer.validated_data['hire_date_to']
#
#             queryset = queryset.filter(**filt)
#
#         return queryset
#
#     @action(detail=False, methods=['post'])
#     def import_employees(self, request):
#         """Import d'employ√©s depuis un fichier CSV/XLSX"""
#         tenant = get_current_tenant_from_request(request)
#         if not tenant:
#             return Response({"detail": "Tenant introuvable"}, status=400)
#
#         serializer = EmployeeImportSerializer(data=request.data)
#         if serializer.is_valid():
#             file = serializer.validated_data['file']
#             update_existing = serializer.validated_data['update_existing']
#
#             try:
#                 if file.name.lower().endswith('.xlsx'):
#                     df = pd.read_excel(file)
#                 else:
#                     df = pd.read_csv(file)
#
#                 imported_count = 0
#                 errors = []
#
#                 with transaction.atomic():
#                     for index, row in df.iterrows():
#                         try:
#                             employee_data = {
#                                 'matricule': row.get('matricule'),
#                                 'first_name': row.get('first_name'),
#                                 'last_name': row.get('last_name'),
#                                 'email': row.get('email'),
#                                 'tenant': tenant,
#                             }
#
#                             if not employee_data['matricule'] or not employee_data['email']:
#                                 raise ValueError("matricule et email sont requis")
#
#                             if update_existing:
#                                 Employee.objects.update_or_create(
#                                     matricule=employee_data['matricule'],
#                                     tenant=tenant,
#                                     defaults=employee_data
#                                 )
#                             else:
#                                 Employee.objects.create(**employee_data)
#
#                             imported_count += 1
#
#                         except Exception as e:
#                             errors.append(f"Ligne {index + 2}: {str(e)}")
#
#                 return Response({
#                     "message": f"{imported_count} employ√©s import√©s",
#                     "errors": errors
#                 })
#
#             except Exception as e:
#                 return Response(
#                     {"error": f"Erreur lors de l'import: {str(e)}"},
#                     status=status.HTTP_400_BAD_REQUEST
#                 )
#
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
#
#     @action(detail=False, methods=['post'])
#     def export_employees(self, request):
#         """Export d'employ√©s"""
#         tenant_id = request.headers.get("X-Tenant-Id")
#         serializer = EmployeeExportSerializer(data=request.data)
#         if serializer.is_valid():
#             export_format = serializer.validated_data['format']
#             fields = serializer.validated_data['fields']
#             filters_data = serializer.validated_data.get('filters', {})
#
#             export_service = EmployeeExportService()
#             result = export_service.export_employees(
#                 tenant_id=tenant_id,
#                 export_format=export_format,
#                 fields=fields,
#                 filters=filters_data
#             )
#
#             if result.get('success'):
#                 return HttpResponse(
#                     result['content'],
#                     content_type=result['content_type'],
#                     headers={'Content-Disposition': f'attachment; filename="{result["filename"]}"'}
#                 )
#             else:
#                 return Response(
#                     {"error": result.get('error', "Export √©chou√©")},
#                     status=status.HTTP_400_BAD_REQUEST
#                 )
#
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
#
#     @action(detail=True, methods=['get'])
#     def leave_balance(self, request, pk=None):
#         """Solde de cong√©s d'un employ√©"""
#         employee = self.get_object()
#         balances = LeaveBalance.objects.filter(employee=employee)
#         serializer = LeaveBalanceSerializer(balances, many=True)
#         return Response(serializer.data)
#
#     @action(detail=True, methods=['get'])
#     def attendance(self, request, pk=None):
#         """Pointage d'un employ√©"""
#         employee = self.get_object()
#         month = int(request.query_params.get('month', timezone.now().month))
#         year = int(request.query_params.get('year', timezone.now().year))
#
#         attendances = Attendance.objects.filter(
#             employee=employee,
#             date__year=year,
#             date__month=month
#         )
#         serializer = AttendanceSerializer(attendances, many=True)
#         return Response(serializer.data)

User = get_user_model()

log = logging.getLogger(__name__)
User = get_user_model()


class EmployeeViewSet(BaseTenantViewSet, viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['first_name', 'last_name', 'email', 'matricule']
    ordering_fields = ['first_name', 'last_name', 'hire_date', 'created_at']

    # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ Utilitaires internes ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    def get_queryset(self):
        qs = Employee.objects.select_related("tenant", "department", "position")

        # ‚úÖ IMPORTANT : si BaseTenantViewSet filtre d√©j√†, garde son comportement.
        # Sinon, applique ton filtre tenant ici (√† adapter √† ton BaseTenantViewSet)
        tenant = getattr(self.request, "tenant", None)
        if tenant is not None:
            qs = qs.filter(tenant=tenant)

        # --- filtres UI ---
        tenant = self.request.query_params.get("tenant")
        department = self.request.query_params.get("department")
        contract_type = self.request.query_params.get("contract_type")
        is_active = self.request.query_params.get("is_active")

        if department:
            qs = qs.filter(department_id=department)

        if contract_type:
            qs = qs.filter(contract_type=contract_type)

        if is_active not in (None, "",):
            # accepte true/false/1/0
            val = str(is_active).lower()
            if val in ("true", "1", "yes"):
                qs = qs.filter(is_active=True)
            elif val in ("false", "0", "no"):
                qs = qs.filter(is_active=False)

        return qs

    def _get_tenant_kwargs(self):
        request = self.request

        # 1) Si un middleware pose d√©j√† request.tenant
        req_tenant = getattr(request, "tenant", None)
        if req_tenant is not None:
            return {"tenant": req_tenant}

        # 2) Header X-Tenant-Id (slug OU UUID)
        raw = request.headers.get("X-Tenant-Id") or request.META.get("HTTP_X_TENANT_ID")
        if raw:
            raw = str(raw).strip()
            tenant = None

            # a) Essayer comme UUID (id)
            try:
                uuid.UUID(raw)
                tenant = Tenant.objects.filter(id=raw, is_active=True).first()
            except ValueError:
                # b) Sinon, on consid√®re que c‚Äôest un slug
                tenant = Tenant.objects.filter(slug=raw, is_active=True).first()

            if tenant is None:
                raise serializers.ValidationError(
                    {"tenant": f"Tenant introuvable pour ¬´ {raw} ¬ª."}
                )

            return {"tenant": tenant}

        # 3) Fallback: tenant li√© √† l'utilisateur (user.employee)
        user = request.user
        if user and not user.is_anonymous:
            emp = getattr(user, "employee", None)
            if emp and emp.tenant_id:
                return {"tenant_id": emp.tenant_id}

        # 4) Rien trouv√© ‚Üí 400 propre
        raise serializers.ValidationError(
            {"tenant": "Impossible de d√©terminer le tenant pour cette requ√™te."}
        )

    def _get_or_create_user_for_employee(self, validated_data):
        """
        Cr√©e ou r√©cup√®re un user √† partir de l'email de l'employ√©.
        Compatible avec custom User sans `username`.
        """
        email = validated_data.get("email")
        first_name = validated_data.get("first_name") or ""
        last_name = validated_data.get("last_name") or ""

        if not email:
            return None

        defaults = {
            "first_name": first_name,
            "last_name": last_name,
            "is_active": True,
        }

        # Ajout de username uniquement si le mod√®le User a ce champ
        if any(f.name == "username" for f in User._meta.get_fields()):
            defaults["username"] = email

        try:
            user, created = User.objects.get_or_create(
                email=email,
                defaults=defaults,
            )
            log.debug(
                "[EmployeeViewSet] user_account=%s (created=%s) pour email=%s",
                user, created, email
            )
        except User.MultipleObjectsReturned:
            user = User.objects.filter(email=email).order_by("id").first()
            log.warning(
                "[EmployeeViewSet] Multiple User pour email=%s, on prend le premier id=%s",
                email, user.id if user else None
            )

        return user

    # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ Hooks DRF ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    def _resolve_tenant(self):
        request = self.request

        # 1) Middleware
        req_tenant = getattr(request, "tenant", None)
        if req_tenant is not None:
            return req_tenant

        # 2) Header X-Tenant-Id (slug ou UUID)
        raw = request.headers.get("X-Tenant-Id") or request.META.get("HTTP_X_TENANT_ID")
        if raw:
            raw = str(raw).strip()
            tenant = None
            try:
                uuid.UUID(raw)
                tenant = Tenant.objects.filter(id=raw, is_active=True).first()
            except ValueError:
                tenant = Tenant.objects.filter(slug=raw, is_active=True).first()

            if tenant is None:
                raise serializers.ValidationError(
                    {"tenant": f"Tenant introuvable pour ¬´ {raw} ¬ª."}
                )
            return tenant

        # 3) fallback user.employee
        user = request.user
        if user and not user.is_anonymous:
            emp = getattr(user, "employee", None)
            if emp and emp.tenant_id:
                return emp.tenant

        raise serializers.ValidationError(
            {"tenant": "Impossible de d√©terminer le tenant pour cette requ√™te."}
        )

    def perform_create(self, serializer):
        tenant = self._resolve_tenant()  # ‚¨ÖÔ∏è objet Tenant
        user = self._get_or_create_user_for_employee(serializer.validated_data)

        extra_kwargs = {"tenant": tenant}
        if user is not None:
            extra_kwargs["user_account"] = user

        with transaction.atomic():
            serializer.save(**extra_kwargs)


class LeaveRequestViewSet(BaseTenantViewSet, viewsets.ModelViewSet):
    queryset = LeaveRequest.objects.all()
    serializer_class = LeaveRequestSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    ordering_fields = ['requested_at', 'start_date']

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtrage avanc√©
        filter_serializer = LeaveRequestFilterSerializer(data=self.request.query_params)
        if filter_serializer.is_valid():
            filt: Dict[str, Any] = {}
            if filter_serializer.validated_data.get('status'):
                filt['status'] = filter_serializer.validated_data['status']
            if filter_serializer.validated_data.get('employee'):
                filt['employee__id'] = filter_serializer.validated_data['employee']
            if filter_serializer.validated_data.get('leave_type'):
                filt['leave_type__id'] = filter_serializer.validated_data['leave_type']
            if filter_serializer.validated_data.get('start_date_from'):
                filt['start_date__gte'] = filter_serializer.validated_data['start_date_from']
            if filter_serializer.validated_data.get('start_date_to'):
                filt['start_date__lte'] = filter_serializer.validated_data['start_date_to']

            queryset = queryset.filter(**filt)

        return queryset

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        obj = self.get_object()
        obj.status = "approved"
        obj.approved_by = getattr(request.user, "employee_profile", None)
        obj.approved_at = timezone.now()
        obj.save()
        return Response({"status": obj.status})

    @action(detail=True, methods=["post"])
    def reject(self, request, pk=None):
        obj = self.get_object()
        obj.status = "rejected"
        obj.rejection_reason = request.data.get('reason', '')
        obj.approved_by = getattr(request.user, "employee_profile", None)
        obj.approved_at = timezone.now()
        obj.save()
        return Response({"status": obj.status})


class PositionViewSet(BaseTenantViewSet, viewsets.ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['title', 'code']
    ordering_fields = ['title', 'created_at']


class LeaveTypeViewSet(BaseTenantViewSet, viewsets.ModelViewSet):
    queryset = LeaveType.objects.all()
    serializer_class = LeaveTypeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'code']
    ordering_fields = ['name', 'created_at']


class AttendanceViewSet(BaseTenantViewSet, viewsets.ModelViewSet):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['post'])
    def bulk_check_in(self, request):
        """
        Pointage massif: body = { "employee_ids": [uuid, ...], "time": "08:30" (optionnel, HH:MM) }
        """
        tenant_id = request.headers.get("X-Tenant-Id")
        employee_ids = request.data.get("employee_ids", [])
        time_str = request.data.get("time")  # "HH:MM"
        if not isinstance(employee_ids, list) or not employee_ids:
            return Response({"detail": "employee_ids requis (liste)"}, status=400)

        today = timezone.localdate()
        check_in_time = None
        if time_str:
            try:
                hh, mm = [int(x) for x in time_str.split(":")]
                check_in_time = timezone.datetime(today.year, today.month, today.day, hh, mm).time()
            except Exception:
                return Response({"detail": "format de time invalide (HH:MM)"}, status=400)

        tenant = get_current_tenant_from_request(request)
        employees = Employee.objects.filter(id__in=employee_ids, tenant=tenant)
        updated = 0
        with transaction.atomic():
            for emp in employees:
                att, _ = Attendance.objects.get_or_create(
                    employee=emp, date=today, defaults={"tenant_id": tenant_id}
                )
                if not att.check_in:
                    att.check_in = check_in_time or timezone.now().time()
                att.status = att.status or 'PRESENT'
                att.save()
                updated += 1
        return Response({"updated": updated})


# -----------------------------
# Recrutement
# -----------------------------

class RecruitmentViewSet(BaseTenantViewSet, viewsets.ModelViewSet):
    queryset = Recruitment.objects.all()  # <-- IMPORTANT (BaseTenantViewSet l‚Äôutilise)
    serializer_class = RecruitmentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['title', 'reference']
    ordering_fields = ['created_at', 'publication_date', 'title']

    def get_queryset(self):
        qs = super().get_queryset().select_related(
            "department", "position", "hiring_manager"
        )

        request = self.request
        status_ = request.query_params.get("status")
        department = request.query_params.get("department")
        position = request.query_params.get("position")
        pub_from = request.query_params.get("publication_date_from")
        pub_to = request.query_params.get("publication_date_to")

        if status_ and status_ != "all":
            qs = qs.filter(status=status_)

        if department:
            qs = qs.filter(department_id=department)

        if position:
            qs = qs.filter(position_id=position)

        if pub_from:
            try:
                d = datetime.fromisoformat(pub_from).date()
                qs = qs.filter(publication_date__gte=d)
            except ValueError:
                pass

        if pub_to:
            try:
                d = datetime.fromisoformat(pub_to).date()
                qs = qs.filter(publication_date__lte=d)
            except ValueError:
                pass

        return qs


class RecruitmentStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def _get_tenant(self, request):
        tenant = getattr(request, "tenant", None)
        if tenant:
            return tenant

        raw = request.headers.get("X-Tenant-Id") or request.META.get("HTTP_X_TENANT_ID")
        if not raw:
            return None

        raw = str(raw).strip()

        # slug d'abord
        t = Tenant.objects.filter(slug=raw, is_active=True).first()
        if t:
            return t

        # uuid ensuite
        try:
            uuid.UUID(raw)
        except ValueError:
            raise ValidationError({"tenant": f"'{raw}' n'est ni un slug, ni un UUID valide."})

        t = Tenant.objects.filter(id=raw, is_active=True).first()
        if not t:
            return None
        return t

    def get(self, request, *args, **kwargs):
        tenant = self._get_tenant(request)
        if not tenant:
            # pas de tenant = pas de data, mais pas 500
            return Response(
                {
                    "total_recruitments": 0,
                    "applications_by_status": {},
                }
            )

        recr_qs = Recruitment.objects.filter(tenant=tenant)

        total_recruitments = recr_qs.count()

        # agr√©gation par statut de candidatures
        apps_qs = JobApplication.objects.filter(recruitment__in=recr_qs)
        by_status = (
            apps_qs.values("status")
            .annotate(total=models.Count("id"))
            .order_by()
        )
        applications_by_status = {
            row["status"]: row["total"] for row in by_status
        }

        return Response(
            {
                "total_recruitments": total_recruitments,
                "applications_by_status": applications_by_status,
            }
        )


# -----------------------------
# Candidatures (fusion des deux d√©finitions)
# -----------------------------
class JobApplicationViewSet(BaseTenantViewSet, viewsets.ModelViewSet):
    queryset = JobApplication.objects.all()
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, JSONParser]

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return JobApplicationDetailSerializer
        return JobApplicationSerializer

    @action(detail=True, methods=["post"], url_path="process_ai")
    def process_ai(self, request, pk=None):
        """
        Lance le traitement IA pour cette candidature et retourne le AIProcessingResult.
        """
        tenant_id = request.headers.get("X-Tenant-Id")
        app = get_object_or_404(JobApplication, pk=pk, tenant_id=tenant_id)

        # Idempotence simple : si d√©j√† COMPLETED et pas de for√ßage ‚Üí on renvoie le dernier r√©sultat
        force = request.query_params.get("force", "false").lower() == "true"
        existing = getattr(app, "ai_processing", None)
        if existing and existing.status == "COMPLETED" and not force:
            return Response(AIProcessingResultSerializer(existing).data)

        service = AIRecruitmentService()
        result = service.process_application(app)
        return Response(AIProcessingResultSerializer(result).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def change_status(self, request, pk=None):
        """Changer le statut d'une candidature"""
        application = self.get_object()
        new_status = request.data.get('status')

        valid_statuses = dict(JobApplication.STATUS_CHOICES).keys()
        if new_status in valid_statuses:
            application.status = new_status
            application.save()
            return Response({"status": application.status})

        return Response(
            {"error": "Statut invalide"},
            status=status.HTTP_400_BAD_REQUEST
        )

    @action(detail=True, methods=['post'])
    def schedule_interview(self, request, pk=None):
        """Planifier un entretien"""
        application = self.get_object()

        interview_data = {
            'job_application': str(application.id),
            'interview_type': request.data.get('interview_type', 'HR'),
            'scheduled_date': request.data.get('scheduled_date'),
            'duration': request.data.get('duration', 60),
            'interviewers': request.data.get('interviewers', []),
            'tenant_id': application.tenant_id,
        }

        serializer = InterviewSerializer(data=interview_data)
        if serializer.is_valid():
            interview = serializer.save()
            return Response(InterviewSerializer(interview).data)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class InterviewViewSet(BaseTenantViewSet, viewsets.ModelViewSet):
    queryset = Interview.objects.all()
    serializer_class = InterviewSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Marquer un entretien comme termin√©"""
        interview = self.get_object()

        interview.conducted_at = timezone.now()
        interview.status = 'COMPLETED'
        # feedback structur√© json ; overall_rating/recommendation/notes peuvent √™tre fournis
        interview.interviewer_feedback = request.data.get('feedback', {})
        interview.overall_rating = request.data.get('overall_rating')
        interview.recommendation = request.data.get('recommendation', '')
        interview.notes = request.data.get('notes', '')

        interview.save()
        return Response(InterviewSerializer(interview).data)


class PerformanceReviewViewSet(BaseTenantViewSet, viewsets.ModelViewSet):
    queryset = PerformanceReview.objects.all()
    serializer_class = PerformanceReviewSerializer
    permission_classes = [IsAuthenticated, HasRHAccess]

    @action(detail=True, methods=['post'])
    def finalize(self, request, pk=None):
        """Finaliser une √©valuation"""
        review = self.get_object()
        review.status = 'FINALIZED'
        review.save()
        return Response({"status": review.status})


# ---------- Contrats ----------
class ContractTypeViewSet(BaseTenantViewSet):
    queryset = ContractType.objects.all()
    serializer_class = ContractTypeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name", "code"]
    ordering_fields = ["name", "created_at"]


EXPORT_FIELD_MAP = {
    # champs simples
    "contract_number": ("contract_number", "N¬∞ Contrat"),
    "title": ("title", "Poste"),
    "status": ("status", "Statut"),
    "start_date": ("start_date", "D√©but"),
    "end_date": ("end_date", "Fin"),
    "base_salary": ("base_salary", "Salaire"),
    "salary_currency": ("salary_currency", "Devise"),
    "weekly_hours": ("weekly_hours", "Heures/sem"),
    "remote_allowed": ("remote_allowed", "T√©l√©travail"),
    "work_location": ("work_location", "Lieu"),

    # relations (valeurs calcul√©es)
    "employee": ("employee__id", "Employ√© (ID)"),
    "employee_name": (None, "Employ√©"),
    "department": ("department__name", "D√©partement"),
    "position": ("position__title", "Poste (position)"),
    "contract_type": ("contract_type__name", "Type contrat"),
    "approved_by": ("approved_by__id", "Approuv√© par (ID)"),
}


def _bool_to_fr(v):
    return "Oui" if v else "Non"


def _safe_str(v):
    if v is None:
        return ""
    return str(v)


def _format_date(d):
    if not d:
        return ""
    # YYYY-MM-DD (stable pour export)
    return d.isoformat()


class EmploymentContractViewSet(BaseTenantViewSet):
    queryset = EmploymentContract.objects.all()
    serializer_class = EmploymentContractSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["contract_number", "title", "employee__email"]
    ordering_fields = ["start_date", "end_date", "created_at"]

    def _base_queryset_for_export(self, request):
        qs = EmploymentContract.objects.select_related(
            "employee", "department", "position", "contract_type", "approved_by"
        )

        # ‚úÖ Multi-tenant
        tenant_id = getattr(request, "tenant_id", None) or request.headers.get("X-Tenant-Id")
        if tenant_id:
            qs = qs.filter(tenant_id=tenant_id)

        return qs

    def _apply_export_filters(self, qs, filters: dict, search: str, ordering: str):
        # filters attendus (exemples): status, contract_type, department, employee, active_only, date ranges...
        if not filters:
            filters = {}

        # Exemples de filtres
        if filters.get("status"):
            qs = qs.filter(status=filters["status"])

        if filters.get("department"):
            qs = qs.filter(department_id=filters["department"])

        if filters.get("contract_type"):
            qs = qs.filter(contract_type_id=filters["contract_type"])

        if filters.get("employee"):
            qs = qs.filter(employee_id=filters["employee"])

        # date range
        if filters.get("start_date_from"):
            qs = qs.filter(start_date__gte=filters["start_date_from"])
        if filters.get("start_date_to"):
            qs = qs.filter(start_date__lte=filters["start_date_to"])

        # search (simple)
        if search:
            s = search.strip()
            qs = qs.filter(
                Q(contract_number__icontains=s) |
                Q(title__icontains=s) |
                Q(employee__first_name__icontains=s) |
                Q(employee__last_name__icontains=s)
            )

        # ordering (si tu veux autoriser)
        if ordering:
            qs = qs.order_by(ordering)

        return qs

    def _build_rows(self, qs, fields):
        rows = []
        for c in qs:
            row = []
            for f in fields:
                if f not in EXPORT_FIELD_MAP:
                    row.append("")
                    continue

                key, _label = EXPORT_FIELD_MAP[f]

                if f == "employee_name":
                    row.append(_safe_str(f"{c.employee.first_name} {c.employee.last_name}".strip()))
                elif f == "remote_allowed":
                    row.append(_bool_to_fr(bool(c.remote_allowed)))
                elif f in ("start_date", "end_date"):
                    row.append(_format_date(getattr(c, f)))
                elif key:
                    # key de type "department__name" => on lit sur l'objet via relations d√©j√† select_related
                    # comme on a l'objet complet, on pr√©f√®re l'acc√®s direct:
                    if f == "department":
                        row.append(_safe_str(c.department.name if c.department else ""))
                    elif f == "position":
                        row.append(_safe_str(c.position.title if c.position else ""))
                    elif f == "contract_type":
                        row.append(_safe_str(c.contract_type.name if c.contract_type else ""))
                    else:
                        # champs simples
                        row.append(_safe_str(getattr(c, key.split("__")[0], "")))
                else:
                    row.append("")
            rows.append(row)
        return rows

    def _export_xlsx(self, filename, headers, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contrats"

        ws.append(headers)
        for r in rows:
            ws.append(r)

        # autosize basique
        for col_idx, _ in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(12, min(45, len(str(headers[col_idx - 1])) + 6))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        return resp

    def _export_csv(self, filename, headers, rows):
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";")
        writer.writerow(headers)
        writer.writerows(rows)

        resp = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
        return resp

    @action(detail=False, methods=["POST"], url_path="export_contracts")
    def export_contracts(self, request, *args, **kwargs):
        ser = EmploymentContractExportSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        fmt = ser.validated_data["format"]
        fields = ser.validated_data["fields"]
        filters = ser.validated_data.get("filters") or {}
        ordering = ser.validated_data.get("ordering") or ""
        search = ser.validated_data.get("search") or ""

        # headers
        headers = []
        for f in fields:
            if f in EXPORT_FIELD_MAP:
                headers.append(EXPORT_FIELD_MAP[f][1])
            else:
                headers.append(f)

        qs = self._base_queryset_for_export(request)
        qs = self._apply_export_filters(qs, filters, search, ordering)

        rows = self._build_rows(qs, fields)

        stamp = timezone.now().strftime("%Y%m%d-%H%M")
        filename = f"contracts_export_{stamp}"

        if fmt == "csv":
            return self._export_csv(filename, headers, rows)

        return self._export_xlsx(filename, headers, rows)

    @action(detail=True, methods=["post"])
    def activate(self, request, pk=None):
        obj = self.get_object()
        obj.status = "ACTIVE"
        obj.save(update_fields=["status"])
        return Response({"status": obj.status})


class ContractAmendmentViewSet(BaseTenantViewSet):
    queryset = ContractAmendment.objects.all()
    serializer_class = ContractAmendmentSerializer
    permission_classes = [IsAuthenticated]


class ContractTemplateViewSet(BaseTenantViewSet):
    queryset = ContractTemplate.objects.all()
    serializer_class = ContractTemplateSerializer
    permission_classes = [IsAuthenticated]


class ContractAlertViewSet(BaseTenantViewSet):
    queryset = ContractAlert.objects.all()
    serializer_class = ContractAlertSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ["due_date", "priority", "created_at"]


class ContractHistoryViewSet(BaseTenantViewSet):
    queryset = ContractHistory.objects.all()
    serializer_class = ContractHistorySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ["performed_at"]


# ---------- RH core ----------
class SalaryHistoryViewSet(BaseTenantViewSet):
    queryset = SalaryHistory.objects.all()
    serializer_class = SalaryHistorySerializer
    permission_classes = [IsAuthenticated]


class HRDocumentViewSet(BaseTenantViewSet):
    queryset = HRDocument.objects.all()
    serializer_class = HRDocumentSerializer
    permission_classes = [IsAuthenticated]


# ---------- Cong√©s avanc√© ----------
class LeaveBalanceViewSet(BaseTenantViewSet):
    queryset = LeaveBalance.objects.all()
    serializer_class = LeaveBalanceSerializer
    permission_classes = [IsAuthenticated]


class LeaveApprovalStepViewSet(BaseTenantViewSet):
    queryset = LeaveApprovalStep.objects.all()
    serializer_class = LeaveApprovalStepSerializer
    permission_classes = [IsAuthenticated]


# ---------- Calendrier & horaires ----------
class HolidayCalendarViewSet(BaseTenantViewSet):
    queryset = HolidayCalendar.objects.all()
    serializer_class = HolidayCalendarSerializer
    permission_classes = [IsAuthenticated]


class HolidayViewSet(BaseTenantViewSet):
    queryset = Holiday.objects.all()
    serializer_class = HolidaySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ["date"]


class WorkScheduleTemplateViewSet(BaseTenantViewSet):
    queryset = WorkScheduleTemplate.objects.all()
    serializer_class = WorkScheduleTemplateSerializer
    permission_classes = [IsAuthenticated]


# ---------- M√©dical ----------
class MedicalRecordViewSet(BaseTenantViewSet):
    queryset = MedicalRecord.objects.all()
    serializer_class = MedicalRecordSerializer
    permission_classes = [IsAuthenticated]


class MedicalVisitViewSet(BaseTenantViewSet):
    queryset = MedicalVisit.objects.all()
    serializer_class = MedicalVisitSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ["visit_date", "created_at"]


class MedicalRestrictionViewSet(BaseTenantViewSet):
    queryset = MedicalRestriction.objects.all()
    serializer_class = MedicalRestrictionSerializer
    permission_classes = [IsAuthenticated]


# ---------- Paie ----------
class PayrollViewSet(BaseTenantViewSet):
    queryset = Payroll.objects.all()
    serializer_class = PayrollSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["payroll_number", "employee__email"]
    ordering_fields = ["period_start", "pay_date", "created_at"]


# ---------- Recrutement avanc√© ----------
class RecruitmentAnalyticsViewSet(BaseTenantViewSet):
    queryset = RecruitmentAnalytics.objects.all()
    serializer_class = RecruitmentAnalyticsSerializer
    permission_classes = [IsAuthenticated]


class JobOfferViewSet(BaseTenantViewSet):
    queryset = JobOffer.objects.all()
    serializer_class = JobOfferSerializer
    permission_classes = [IsAuthenticated]


class RecruitmentWorkflowViewSet(BaseTenantViewSet):
    queryset = RecruitmentWorkflow.objects.all()
    serializer_class = RecruitmentWorkflowSerializer
    permission_classes = [IsAuthenticated]


class InterviewFeedbackViewSet(BaseTenantViewSet):
    queryset = InterviewFeedback.objects.all()
    serializer_class = InterviewFeedbackSerializer
    permission_classes = [IsAuthenticated]
